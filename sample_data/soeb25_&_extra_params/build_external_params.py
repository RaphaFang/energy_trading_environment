"""
Build the external-parameter files.

Three outputs, split by source so no row carries columns that don't apply:

  soeb25_params.csv             One publication, one URL, every row verified.
                                Source columns are table + cell coordinates.
  dk_tax_and_tariff_params.csv  Danish tax law and company tariffs. Every row has
                                a different source, so source columns are per-row.
  gaps.csv                      What is still missing. Different shape entirely -
                                a gap has no value or unit, it has a place to look.

Everything in soeb25_params.csv is read programmatically from the workbook.
No hand-typed numbers, no currency conversion, no deflation - those are
modelling decisions and belong downstream.

Re-run against a later SOEB edition by changing XLSX, then diff the CSV.
"""
import csv
import openpyxl

XLSX = "/mnt/user-data/uploads/SØB25_webudgave_marts26.xlsx"
OUTDIR = "/mnt/user-data/outputs"

# Cited once here rather than repeated on all 280 rows of soeb25_params.csv:
PUBLICATION = ("Energistyrelsen, Samfundsokonomiske beregningsforudsaetninger 2025 "
               "(SOEB25), webudgave marts 2026")
PUBLICATION_URL = "https://ens.dk/analyser-og-statistik/samfundsoekonomiske-analysemetoder"
DATE = "2026-08-09"


def col_letter(i):
    s = ""
    while i > 0:
        i, r = divmod(i - 1, 26)
        s = chr(65 + r) + s
    return s


def write(path, fields, rows):
    with open(f"{OUTDIR}/{path}", "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(rows)
    print(f"{len(rows):4d} rows -> {path}")


def build_soeb():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    rows = []

    def add(param, year, value, unit, table, cell, note):
        rows.append({"param": param, "year": year, "value": value, "unit": unit,
                     "source_table": table, "source_cell": cell, "note": note})

    # Tabel 1: heating values
    ws = wb["Tabel 1"]
    hv = {"Naturgas": ("heating_value_naturgas", "GJ/1000Nm3"),
          "Halm (15 % vandindhold)": ("heating_value_halm", "GJ/ton"),
          "Skovflis (Nåletræ, 40 % vandindhold)": ("heating_value_skovflis", "GJ/ton"),
          "Træpiller (7 % vandindhold)": ("heating_value_traepiller", "GJ/ton"),
          "Affald": ("heating_value_affald", "GJ/ton")}
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 1).value in hv:
            name, unit = hv[ws.cell(r, 1).value]
            add(name, "", ws.cell(r, 2).value, unit, "Tabel 1", f"B{r}",
                "Nedre braendvaerdi (lower heating value).")

    # Tabel 1: net price index, 2025 = 1
    for r in range(1, ws.max_row + 1):
        y = ws.cell(r, 8).value
        if isinstance(y, int) and 2015 <= y <= 2050:
            add("price_index_2025base", y, ws.cell(r, 9).value, "index",
                "Tabel 1", f"I{r}",
                "FM nettoprisindeks. Use to put values from different SOEB "
                "editions on one price basis.")

    # Tabel 2: fuel prices an forbrugssted
    ws = wb["Tabel 2"]
    want = {("An kraftværk", "Halm"): "fuel_price_halm_an_kraftvaerk",
            ("An kraftværk", "Træflis"): "fuel_price_traeflis_an_kraftvaerk",
            ("An kraftværk", "Træpiller (industri)"):
                "fuel_price_traepiller_industri_an_kraftvaerk",
            ("An værk", "Halm"): "fuel_price_halm_an_vaerk",
            ("An værk", "Træflis"): "fuel_price_traeflis_an_vaerk",
            ("An værk", "Træpiller (industri)"):
                "fuel_price_traepiller_industri_an_vaerk"}
    group, colmap = None, {}
    for c in range(2, ws.max_column + 1):
        g = ws.cell(3, c).value
        if g:
            group = g.strip()
        fuel = ws.cell(4, c).value
        if fuel and (group, fuel.strip()) in want:
            colmap[c] = want[(group, fuel.strip())]
    for r in range(5, ws.max_row + 1):
        y = ws.cell(r, 1).value
        if isinstance(y, int):
            for c, name in colmap.items():
                add(name, y, ws.cell(r, c).value, "DKK2025/GJ", "Tabel 2",
                    f"{col_letter(c)}{r}",
                    "Faktorpris, excl. taxes/subsidies/VAT. An kraftvaerk = "
                    "central CHP. An vaerk = decentral CHP / DH plant.")

    # Tabel 12: emission coefficients
    ws = wb["Tabel 12"]
    ef = {("Affald", "Dampturbine"): "ef_co2_affald",
          ("Halm*", "Dampturbine"): "ef_co2_halm",
          ("Træ* (eks. træpiller)", "Dampturbine"): "ef_co2_trae",
          ("Træpiller", "Dampturbine"): "ef_co2_traepiller",
          ("Kul", "Dampturbine"): "ef_co2_kul"}
    seen = set()
    for r in range(1, ws.max_row + 1):
        key = (ws.cell(r, 1).value, ws.cell(r, 2).value)
        if key in ef and key not in seen:
            seen.add(key)
            add(ef[key], 2022, ws.cell(r, 3).value, "kg CO2/GJ", "Tabel 12",
                f"C{r}",
                "Kilde DCE. Table note: for biomasse og biogas er anvendt en CO2 "
                "emissionsfaktor paa 0 - accounting convention, not physics.")

    # Tabel 13: gas CO2 coefficient by year
    ws = wb["Tabel 13"]
    for r in range(1, ws.max_row + 1):
        y = ws.cell(r, 1).value
        if isinstance(y, int) and 2000 <= y <= 2050:
            add("ef_co2_ledningsgas", y, ws.cell(r, 2).value, "kg CO2/GJ",
                "Tabel 13", f"B{r}",
                "57.1 for 2025-2031, then 0 from 2032 on marginal-biogas logic. "
                "For a 2019-2026 historical study use 57.1.")

    # Tabel 10: electricity transport + margin, largest consumer band
    ws = wb["Tabel 10"]
    for r in range(1, ws.max_row + 1):
        y = ws.cell(r, 1).value
        if isinstance(y, int) and 2000 <= y <= 2050:
            add("el_transport_margin_over_70000MWh", y, ws.cell(r, 8).value,
                "DKK2025/MWh", "Tabel 10", f"H{r}",
                "NOT refundable, unlike elafgift. Candidate explanation for why "
                "real DK2 power-to-heat runs at only 1.4% of production.")

    # Tabel 15: CO2 price
    ws = wb["Tabel 15"]
    for r in range(1, ws.max_row + 1):
        y = ws.cell(r, 1).value
        if isinstance(y, int) and 2000 <= y <= 2050:
            add("co2_price_soeb_central", y, ws.cell(r, 2).value, "DKK2025/ton",
                "Tabel 15", f"B{r}",
                "Faktorpris. Forecast, not observed. Prefer ICAP observed prices "
                "for a historical backtest; use this only as a cross-check.")

    write("soeb25_params.csv",
          ["param", "year", "value", "unit", "source_table", "source_cell", "note"],
          rows)


def build_tax_tariff():
    F = ["param", "year", "value", "unit", "source_name", "source_url",
         "source_ref", "retrieved_date", "note"]
    ARC_URL = ("https://a-r-c.dk/erhverv/erhvervsaffald-til-forbraending/"
               "forbraending-priser/")
    rows = [
        {"param": "elafgift_dh_producer_net", "year": "2021-", "value": 0.4,
         "unit": "oere/kWh",
         "source_name": "Skatteforvaltningen, Den juridiske vejledning 2025-1",
         "source_url": "https://info.skat.dk/data.aspx?oid=2444248",
         "source_ref": "E.A.4.3.6.2, Bemaerk", "retrieved_date": DATE,
         "note": "VAT-registered heat producers are refunded the difference "
                 "between the 0.4 oere/kWh minimum and the ordinary rate. ELAL "
                 "para 11 stk.1 and para 11c, as amended by lov 2225 of "
                 "29-12-2020. Applies to DH heat pumps AND electric boilers. The "
                 "household elvarmesats explicitly CANNOT be used by heat "
                 "producers. Refund dates from lov 2225/2020, so 2019-2020 "
                 "differs - see gaps.csv."},
        {"param": "gate_fee_arc_rest_erhverv", "year": 2025, "value": 635.00,
         "unit": "DKK/ton excl VAT",
         "source_name": "ARC Affaldsenergianlaeg, takstblad",
         "source_url": ARC_URL,
         "source_ref": "Takster pr. 1. november 2025", "retrieved_date": DATE,
         "note": "Commercial residual waste. REVENUE to the plant, so it enters "
                 "the model as negative fuel cost. Do not also add a separate "
                 "CO2-tax term without checking for double counting - the tariff "
                 "already covers ARC's own waste taxes."},
        {"param": "gate_fee_arc_rest_erhverv_med_haandtering", "year": 2025,
         "value": 725.00, "unit": "DKK/ton excl VAT",
         "source_name": "ARC Affaldsenergianlaeg, takstblad",
         "source_url": ARC_URL,
         "source_ref": "Takster pr. 1. november 2025", "retrieved_date": DATE,
         "note": "Same stream, with handling."},
        {"param": "gate_fee_arc_b_takst_municipal", "year": 2022, "value": 487,
         "unit": "DKK/ton",
         "source_name": "Koebenhavns Kommune, Borgerrepraesentationen 22-09-2022",
         "source_url": "https://www.kk.dk/dagsordener-og-referater/"
                       "Borgerrepr%C3%A6sentationen/m%C3%B8de-22092022/referat/punkt-10",
         "source_ref": "Punkt 10", "retrieved_date": DATE,
         "note": "STALE. Municipal household-waste stream, different from the "
                 "commercial tariff above. The document itself states the rate "
                 "needed to rise well over 100 kr/ton. Current value unknown."},
    ]
    write("dk_tax_and_tariff_params.csv", F, rows)


def build_gaps():
    F = ["gap_id", "what_is_missing", "why_it_matters", "where_to_look",
         "do_not_use", "blocks"]
    rows = [
        {"gap_id": "biomass_price_2019_2024",
         "what_is_missing": "Delivered biomass prices (halm, traeflis, traepiller) "
                            "for 2019-2024",
         "why_it_matters": "SOEB25 Tabel 2 and Tabel 5 both start at 2025, so a "
                           "2019-2026 backtest has no fuel price for 6 of 8 years. "
                           "Fuel price is the first-order driver of whether a "
                           "biomass CHP runs at all.",
         "where_to_look": "Option 1: stitch the near-year values from older SOEB "
                          "editions (2018, 2019, 2021, 2023), each of which prices "
                          "its own first years off contemporaneous forwards, then "
                          "put them on one basis using price_index_2025base in "
                          "soeb25_params.csv. Option 2: Energistyrelsen "
                          "'Energipriser og afgifter' quarterly observed statistics.",
         "do_not_use": "SOEB25's 2025 figure as a proxy for earlier years. Tabel 5 "
                       "Note 3 states its forwards were drawn January 2025, so even "
                       "the 2025 cell is a forward view, not an outturn.",
         "blocks": "Backtest years 2019-2024, all biomass CHP units"},
        {"gap_id": "co2_afgift_affald",
         "what_is_missing": "Statutory Danish CO2 tax rate on waste incineration",
         "why_it_matters": "Three of the six DK2 agents burn waste; the tax sits "
                           "directly in their marginal cost.",
         "where_to_look": "skat.dk statutory rate schedule under "
                          "kuldioxidafgiftsloven; Groen Skattereform phase-in "
                          "2025-2030.",
         "do_not_use": "News articles - secondary sources disagree on which rate "
                       "applies to ETS-covered incinerators.",
         "blocks": "Marginal cost of ARC, Vestforbraending, ARGO"},
        {"gap_id": "gate_fee_vestforbraending",
         "what_is_missing": "Vestforbraending gate fee",
         "why_it_matters": "Negative fuel cost sets this agent's merit-order "
                           "position.",
         "where_to_look": "Vestforbraending takstblad or annual report.",
         "do_not_use": "ARC's tariff as a proxy - separate cooperatives, separate "
                       "cost bases.",
         "blocks": "Vestforbraending marginal cost"},
        {"gap_id": "gate_fee_argo",
         "what_is_missing": "ARGO gate fee",
         "why_it_matters": "Same as above.",
         "where_to_look": "ARGO takstblad or annual report.",
         "do_not_use": "ARC's tariff as a proxy.",
         "blocks": "ARGO marginal cost"},
        {"gap_id": "elafgift_2019_2020",
         "what_is_missing": "Electricity tax regime for DH heat producers before "
                            "lov 2225 of 29-12-2020",
         "why_it_matters": "The 0.4 oere/kWh refund only applies from 2021, but "
                           "the backtest starts in 2019.",
         "where_to_look": "Historical elpatronordning provisions in ELAL; "
                          "Skatteforvaltningen historical vejledning editions.",
         "do_not_use": "The current 0.4 oere/kWh figure for 2019-2020.",
         "blocks": "Power-to-heat cost, backtest years 2019-2020"},
        {"gap_id": "dk2_unit_capacities",
         "what_is_missing": "Actual installed MW_e and MW_th per DK2 agent",
         "why_it_matters": "The DEA catalogue gives prototype sizes, not the real "
                           "fleet.",
         "where_to_look": "Energinet CapacityPerMunicipality (already in duckdb); "
                          "ENTSO-E production units list; plant annual reports.",
         "do_not_use": "Catalogue prototype capacities as if they were real units.",
         "blocks": "All six agents"},
    ]
    write("gaps.csv", F, rows)


if __name__ == "__main__":
    print(f"source: {PUBLICATION}\n        {PUBLICATION_URL}\n")
    build_soeb()
    build_tax_tariff()
    build_gaps()
